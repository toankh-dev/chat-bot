from typing import List, Dict, Any
from datetime import datetime
from langchain.text_splitter import (
    RecursiveCharacterTextSplitter,
    TokenTextSplitter,
    MarkdownTextSplitter,
)
from openpyxl import load_workbook
from pathlib import Path


# ==========================
# 1️⃣ SPECIFIC CHUNKER BY SOURCE
# ==========================

def chunk_slack(messages: List[Dict[str, Any]], window_size=10, overlap=2) -> List[Dict]:
    """Chunk Slack messages by conversation window"""
    chunks = []
    for i in range(0, len(messages), window_size - overlap):
        group = messages[i:i + window_size]
        text = "\n".join([f"{m['user']}: {m['text']}" for m in group if m.get("text")])
        if not text.strip():
            continue
        metadata = {
            "source": "slack",
            "channel": group[0].get("channel", "unknown"),
            "start_time": group[0].get("timestamp"),
            "end_time": group[-1].get("timestamp"),
            "message_count": len(group),
        }
        chunks.append({"text": text, "metadata": metadata})
    return chunks


def chunk_code(code_text: str, language: str = "python") -> List[Dict]:
    """Chunk code by function/class boundaries"""
    # Fallback to RecursiveCharacterTextSplitter with code-friendly separators
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=1200,
        chunk_overlap=200,
        separators=["\nclass ", "\ndef ", "\n\n", "\n", " ", ""]
    )
    sections = splitter.split_text(code_text)
    return [{"text": s, "metadata": {"source": "code", "language": language}} for s in sections]


def chunk_markdown(markdown_text: str) -> List[Dict]:
    """Chunk markdown or wiki"""
    splitter = MarkdownTextSplitter(chunk_size=1200, chunk_overlap=200)
    sections = splitter.split_text(markdown_text)
    return [{"text": s, "metadata": {"source": "markdown"}} for s in sections]


def chunk_backlog_issues(issues: List[Dict[str, Any]]) -> List[Dict]:
    """Chunk backlog / Jira issue"""
    splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=150)
    chunks = []
    for issue in issues:
        text = f"Title: {issue.get('title', '')}\nDescription: {issue.get('description', '')}"
        comments = issue.get("comments", [])
        if comments:
            text += "\n\nComments:\n" + "\n".join(comments)
        split_parts = splitter.split_text(text)
        for part in split_parts:
            chunks.append({
                "text": part,
                "metadata": {
                    "source": "backlog",
                    "issue_id": issue.get("id"),
                    "project": issue.get("project"),
                    "created_at": issue.get("created_at"),
                }
            })
    return chunks


def chunk_excel(file_path: str) -> List[Dict]:
    """
    Improved Excel chunking:
    - Reads both structured (DataFrame) and unstructured (cell-by-cell) content.
    - Each row is a chunk if tabular data is detected.
    - Full sheet text is also added as fallback.
    """
    documents = []
    file_name = Path(file_path).name

    # Load workbook fully (for non-tabular data)
    wb = load_workbook(file_path, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # --- Step 1: Extract raw text of entire sheet (cell-by-cell) ---
        raw_sheet_text = []
        for row in ws.iter_rows(values_only=True):
            row_text = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if row_text:
                raw_sheet_text.append(" | ".join(row_text))

        # Create a full-sheet text version (for context embedding)
        if raw_sheet_text:
            full_text = f"Sheet: {sheet_name}\n\n" + "\n".join(raw_sheet_text)
            splitter = RecursiveCharacterTextSplitter(chunk_size=2000, chunk_overlap=200)
            for i, chunk in enumerate(splitter.split_text(full_text)):
                documents.append({
                    "text": chunk,
                    "metadata": {
                        "source": "excel",
                        "file": file_name,
                        "sheet": sheet_name,
                        "type": "full_sheet",
                        "chunk_index": i
                    }
                })

        # --- Step 2: Try structured read using pandas ---
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        except Exception:
            df = None

        if df is not None and not df.empty:
            # Individual row chunks
            for idx, row in df.iterrows():
                row_parts = []
                for col in df.columns:
                    val = row[col]
                    if pd.notna(val) and str(val).strip():
                        clean_col = col if not str(col).startswith("Unnamed:") else f"Column_{col.split(':')[-1].strip()}"
                        row_parts.append(f"{clean_col}: {val}")
                if row_parts:
                    row_text = " | ".join(row_parts)
                    documents.append({
                        "text": row_text,
                        "metadata": {
                            "source": "excel",
                            "file": file_name,
                            "sheet": sheet_name,
                            "row_number": int(idx) + 2,
                            "total_rows": len(df),
                            "type": "row"
                        }
                    })

            # Sheet summary chunk (based on first 10 rows)
            sample_rows = []
            for idx, row in df.head(10).iterrows():
                row_parts = []
                for col in df.columns:
                    val = row[col]
                    if pd.notna(val) and str(val).strip():
                        clean_col = col if not str(col).startswith("Unnamed:") else f"Column_{col.split(':')[-1].strip()}"
                        row_parts.append(f"{clean_col}: {val}")
                if row_parts:
                    sample_rows.append(" | ".join(row_parts))

            if sample_rows:
                summary_text = (
                    f"Sheet: {sheet_name}\n"
                    f"Total rows: {len(df)}\n"
                    f"Columns: {', '.join(map(str, df.columns[:10]))}\n\n"
                    "Sample (first 10 rows):\n" + "\n".join(sample_rows)
                )
                splitter = RecursiveCharacterTextSplitter(chunk_size=2000, chunk_overlap=200)
                for i, chunk in enumerate(splitter.split_text(summary_text)):
                    documents.append({
                        "text": chunk,
                        "metadata": {
                            "source": "excel",
                            "file": file_name,
                            "sheet": sheet_name,
                            "type": "sheet_summary",
                            "chunk_index": i
                        }
                    })

    return documents


def chunk_commit_or_pr(data: Dict[str, Any]) -> List[Dict]:
    """Chunk Git commit or Pull Request"""
    text = f"Title: {data.get('title')}\n\nDescription:\n{data.get('description', '')}"
    if diff := data.get("diff"):
        text += f"\n\nDiff:\n{diff}"
    splitter = TokenTextSplitter(chunk_size=600, chunk_overlap=100)
    sections = splitter.split_text(text)
    return [
        {"text": s, "metadata": {"source": "git", "repo": data.get("repo"), "type": data.get("type", "commit")}}
        for s in sections
    ]


def chunk_discord(messages: List[Dict[str, Any]], window_size=10, overlap=2) -> List[Dict]:
    """Chunk Discord messages by conversation window"""
    chunks = []
    for i in range(0, len(messages), window_size - overlap):
        group = messages[i:i + window_size]
        text = "\n".join([f"{m.get('author', 'Unknown')}: {m.get('content', '')}" for m in group if m.get("content")])
        if not text.strip():
            continue
        metadata = {
            "source": "discord",
            "channel": group[0].get("channel_id", "unknown"),
            "start_time": group[0].get("timestamp"),
            "end_time": group[-1].get("timestamp"),
            "message_count": len(group),
        }
        chunks.append({"text": text, "metadata": metadata})
    return chunks


# ==========================
# 2️⃣ MAIN ROUTER
# ==========================

def chunk_data(source_type: str, data: Any, **kwargs) -> List[Dict]:
    """
    Router selects appropriate chunking strategy by data type

    Args:
        source_type: Data source type (slack, discord, code, markdown, etc.)
        data: Data to be chunked
        **kwargs: Optional parameters for specific chunker

    Returns:
        List[Dict]: List of chunks with format {"text": str, "metadata": dict}
    """
    source_type = source_type.lower()

    if source_type == "slack":
        return chunk_slack(data,
                          window_size=kwargs.get("window_size", 10),
                          overlap=kwargs.get("overlap", 2))

    elif source_type == "discord":
        return chunk_discord(data,
                           window_size=kwargs.get("window_size", 10),
                           overlap=kwargs.get("overlap", 2))

    elif source_type in ["code", "github_code", "gitlab_code"]:
        language = kwargs.get("language", "python")
        return chunk_code(data, language=language)

    elif source_type in ["markdown", "wiki", "readme"]:
        return chunk_markdown(data)

    elif source_type in ["backlog", "jira"]:
        return chunk_backlog_issues(data)

    elif source_type in ["excel", "spreadsheet"]:
        return chunk_excel(data)

    elif source_type in ["commit", "pr", "pull_request"]:
        return chunk_commit_or_pr(data)

    else:
        # fallback: semantic / text
        splitter = RecursiveCharacterTextSplitter(
            chunk_size=kwargs.get("chunk_size", 1000),
            chunk_overlap=kwargs.get("chunk_overlap", 150)
        )
        parts = splitter.split_text(str(data))
        return [{"text": p, "metadata": {"source": source_type}} for p in parts]


# ==========================
# 3️⃣ QUICK TEST (OPTIONAL)
# ==========================
if __name__ == "__main__":
    # example: chunk Slack messages
    messages = [
        {"user": "Alice", "text": "Hey team, any update on bug #42?", "timestamp": "2025-10-27T10:00", "channel": "dev"},
        {"user": "Bob", "text": "Yes, fixed and merged!", "timestamp": "2025-10-27T10:05", "channel": "dev"},
    ]
    docs = chunk_data("slack", messages)
    print(f"Slack chunks: {len(docs)}")
    print(docs[0]["text"])
